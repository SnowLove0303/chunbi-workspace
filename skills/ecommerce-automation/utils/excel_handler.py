"""
Excel 文件处理工具
"""

import os
import pandas as pd
from datetime import datetime
from loguru import logger


class ExcelHandler:
    """Excel 文件处理器"""

    def __init__(self):
        """初始化 Excel 处理器"""
        pass

    def save_to_excel(self, df, filepath, sheet_name="Sheet1"):
        """
        将 DataFrame 保存到 Excel 文件
        
        Args:
            df: pandas DataFrame
            filepath: 保存路径
            sheet_name: 工作表名称
        """
        try:
            # 确保目录存在
            dir_path = os.path.dirname(filepath)
            if dir_path and not os.path.exists(dir_path):
                os.makedirs(dir_path, exist_ok=True)
            
            # 保存到 Excel
            df.to_excel(filepath, index=False, sheet_name=sheet_name)
            
            logger.info(f"数据已保存到 Excel: {filepath} ({len(df)} 行)")
            
            return True
            
        except Exception as e:
            logger.error(f"保存 Excel 失败：{e}")
            return False

    def read_from_excel(self, filepath, sheet_name=0):
        """
        从 Excel 文件读取数据
        
        Args:
            filepath: 文件路径
            sheet_name: 工作表名称或索引
            
        Returns:
            DataFrame: 读取的数据
        """
        try:
            if not os.path.exists(filepath):
                logger.error(f"文件不存在：{filepath}")
                return None
            
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            logger.info(f"从 Excel 读取数据：{filepath} ({len(df)} 行)")
            
            return df
            
        except Exception as e:
            logger.error(f"读取 Excel 失败：{e}")
            return None

    def append_to_excel(self, df, filepath, sheet_name="Sheet1"):
        """
        追加数据到现有 Excel 文件
        
        Args:
            df: pandas DataFrame
            filepath: 文件路径
            sheet_name: 工作表名称
        """
        try:
            if os.path.exists(filepath):
                # 读取现有数据
                existing_df = self.read_from_excel(filepath, sheet_name)
                
                if existing_df is not None and not existing_df.empty:
                    # 合并数据
                    combined_df = pd.concat([existing_df, df], ignore_index=True)
                else:
                    combined_df = df
            else:
                combined_df = df
            
            # 保存合并后的数据
            self.save_to_excel(combined_df, filepath, sheet_name)
            
            logger.info(f"数据已追加到 Excel: {filepath}")
            
            return True
            
        except Exception as e:
            logger.error(f"追加数据到 Excel 失败：{e}")
            return False

    def merge_excel_files(self, filepaths, output_filepath, sheet_name="Merged"):
        """
        合并多个 Excel 文件
        
        Args:
            filepaths: 文件路径列表
            output_filepath: 输出文件路径
            sheet_name: 工作表名称
        """
        try:
            all_dfs = []
            
            for filepath in filepaths:
                df = self.read_from_excel(filepath)
                if df is not None and not df.empty:
                    all_dfs.append(df)
            
            if all_dfs:
                merged_df = pd.concat(all_dfs, ignore_index=True)
                self.save_to_excel(merged_df, output_filepath, sheet_name)
                
                logger.info(f"已合并 {len(filepaths)} 个文件到：{output_filepath}")
                return True
            else:
                logger.warning("没有可合并的数据")
                return False
                
        except Exception as e:
            logger.error(f"合并 Excel 文件失败：{e}")
            return False

    def create_summary_report(self, data_dict, output_filepath):
        """
        创建汇总报表（多工作表）
        
        Args:
            data_dict: 字典，键为工作表名，值为 DataFrame
            output_filepath: 输出文件路径
        """
        try:
            with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"汇总报表已创建：{output_filepath} ({len(data_dict)} 个工作表)")
            
            return True
            
        except Exception as e:
            logger.error(f"创建汇总报表失败：{e}")
            return False

    def format_currency_columns(self, filepath, columns=None):
        """
        格式化 Excel 中的货币列
        
        Args:
            filepath: 文件路径
            columns: 需要格式化的列名列表
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            if columns is None:
                # 默认格式化包含价格的列
                columns = ["price", "sales", "profit", "cost"]
            
            # 获取列名所在行（第一行）
            header_row = 1
            column_indices = {}
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value and str(cell_value).lower() in [c.lower() for c in columns]:
                    column_indices[str(cell_value)] = col
            
            # 应用货币格式
            for col_name, col_idx in column_indices.items():
                for row in range(header_row + 1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '¥#,##0.00'
            
            wb.save(filepath)
            wb.close()
            
            logger.info(f"已格式化货币列：{filepath}")
            
            return True
            
        except Exception as e:
            logger.error(f"格式化 Excel 失败：{e}")
            return False


if __name__ == "__main__":
    # 测试
    handler = ExcelHandler()
    
    # 创建测试数据
    test_data = {
        "商品名称": ["商品 A", "商品 B", "商品 C"],
        "价格": [10.5, 20.0, 15.8],
        "销量": [100, 200, 150],
        "利润": [5.25, 10.0, 7.9]
    }
    
    df = pd.DataFrame(test_data)
    
    # 保存测试
    test_file = "test_output.xlsx"
    handler.save_to_excel(df, test_file)
    
    # 读取测试
    df_read = handler.read_from_excel(test_file)
    print("\n=== Excel 处理测试 ===\n")
    print(df_read)
    
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f"\n测试文件已清理：{test_file}")
